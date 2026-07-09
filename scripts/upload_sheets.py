import os
import openpyxl
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build

# Load environment variables from .env.local if present
def load_env():
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), '.env.local')
    if os.path.exists(env_path):
        with open(env_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, val = line.split('=', 1)
                    # Strip quotes if present
                    val = val.strip().strip('"').strip("'")
                    os.environ[key.strip()] = val

load_env()

EXCEL_PATH = r"D:\e-com example\E-Commerce-Analysis-main\Excel\Sales Overview by Product.xlsx"
SPREADSHEET_ID = os.getenv("GOOGLE_SPREADSHEET_ID")
SERVICE_ACCOUNT_EMAIL = os.getenv("GOOGLE_SERVICE_ACCOUNT_EMAIL")
PRIVATE_KEY = os.getenv("GOOGLE_PRIVATE_KEY")

TABS_TO_UPLOAD = [
    "Product Database",
    "Purchase Value change",
    "Credit Limit",
    "Sheet3"
]

def main():
    if not SPREADSHEET_ID:
        print("ERROR: GOOGLE_SPREADSHEET_ID is not defined in .env.local.")
        print("Please configure your .env.local file first before running this script.")
        return

    if not PRIVATE_KEY or not SERVICE_ACCOUNT_EMAIL:
        print("ERROR: Google Cloud Service Account credentials are not defined in .env.local.")
        return

    print(f"Reading Excel file: {EXCEL_PATH}...")
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at {EXCEL_PATH}")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # Format the private key from env
    formatted_key = PRIVATE_KEY.replace('\\n', '\n')

    # Authenticate with Google API
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    credentials = Credentials.from_service_account_info({
        "client_email": SERVICE_ACCOUNT_EMAIL,
        "private_key": formatted_key,
        "type": "service_account",
        "project_id": "ecom-example-501902",
        "token_uri": "https://oauth2.googleapis.com/token"
    }, scopes=scopes)

    service = build("sheets", "v4", credentials=credentials)
    sheets_api = service.spreadsheets()

    print("Connected to Google Sheets API successfully.")

    for sheet_name in TABS_TO_UPLOAD:
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in Excel workbook. Skipping...")
            continue

        print(f"\nProcessing tab '{sheet_name}'...")
        excel_sheet = wb[sheet_name]
        
        # Extract rows
        rows = []
        for r in excel_sheet.iter_rows(values_only=True):
            # Convert values to strings or standard types for Google Sheets JSON serialization
            row_data = []
            for cell in r:
                if cell is None:
                    row_data.append("")
                elif hasattr(cell, "strftime"): # handles datetime objects
                    row_data.append(cell.strftime("%Y-%m-%d"))
                else:
                    row_data.append(cell)
            rows.append(row_data)

        # Filter out completely empty rows from the end
        while rows and not any(cell != "" for cell in rows[-1]):
            rows.pop()

        if not rows:
            print(f"Sheet '{sheet_name}' is empty. Skipping...")
            continue

        # Prepare update range
        # Clear the old sheet data first
        try:
            sheets_api.values().clear(
                spreadsheetId=SPREADSHEET_ID,
                range=f"'{sheet_name}'!A:Z"
            ).execute()
        except Exception as e:
            # If the tab does not exist in Google Sheets, we need to create it
            if "not found" in str(e).lower() or "invalid grp" in str(e).lower() or "parse error" in str(e).lower() or "400" in str(e):
                print(f"Creating new tab '{sheet_name}' in Google Sheets...")
                try:
                    sheets_api.batchUpdate(
                        spreadsheetId=SPREADSHEET_ID,
                        body={
                            "requests": [{
                                "addSheet": {
                                    "properties": {
                                        "title": sheet_name
                                    }
                                }
                            }]
                        }
                    ).execute()
                except Exception as create_err:
                    print(f"Failed to create tab '{sheet_name}': {create_err}")
                    continue
            else:
                print(f"Error clearing tab '{sheet_name}': {e}")
                continue

        # Upload new data
        body = {
            "values": rows
        }
        
        print(f"Uploading {len(rows)} rows to tab '{sheet_name}'...")
        try:
            sheets_api.values().update(
                spreadsheetId=SPREADSHEET_ID,
                range=f"'{sheet_name}'!A1",
                valueInputOption="RAW",
                body=body
            ).execute()
            print(f"OK: Uploaded '{sheet_name}' successfully!")
        except Exception as e:
            print(f"Failed to upload '{sheet_name}': {e}")

    print("\nAll done! Your Google Sheet is now fully synchronized with local Excel database.")

if __name__ == "__main__":
    main()
